#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import paramiko
import json
import pandas as pd
from typing import Optional, Dict
import time
import datetime
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ==================== 日志配置 ====================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [%(funcName)s:%(lineno)s] %(message)s"
)
# 抑制 paramiko 的详细日志
logging.getLogger("paramiko").setLevel(logging.WARNING)

# ==================== 巡检配置 ====================
NODES = [
    {"type": "pgsql", "ip": "10.10.0.85", "purpose": "用户、认证中心数据库", "data_dir": "/", "disk_type": "本地", "real_path": ""},
    {"type": "pgsql", "ip": "10.10.0.93", "purpose": "iEIS数据库", "data_dir": "/iEIS-PG/pgdata", "disk_type": "nas", "real_path": "/var/lib/docker/volumes/iEIS-PG/_data"},
    {"type": "pgsql", "ip": "10.10.0.120", "purpose": "mes数据库-固态盘", "data_dir": "/citus/pgdata/", "disk_type": "nas", "real_path": "/var/lib/docker/volumes/postgres_nas-163/_data"},
    {"type": "pgsql", "ip": "10.10.0.120", "purpose": "mes数据库-机械盘", "data_dir": "/volume1/Prod1/pgdata/", "disk_type": "nas", "real_path": "/var/lib/docker/volumes/postgres_nas-vol1/_data"},
    {"type": "mssql", "ip": "10.10.0.39", "purpose": "贴片机取数数据库", "data_dir": "D盘", "disk_type": "本地", "real_path": ""},
    {"type": "mssql", "ip": "10.10.0.43", "purpose": "生产sqlserver服务器", "data_dir": "E盘", "disk_type": "nas", "real_path": ""},
    {"type": "redis", "ip": "10.10.0.140", "purpose": "redis服务器", "data_dir": "/prod/redis-data", "disk_type": "nas", "real_path": "/var/lib/docker/volumes/rmq-redis-etcd_nas-redis-163/_data"},
    {"type": "rocketmq", "ip": "10.10.0.140", "purpose": "rocketmq服务器", "data_dir": "/prod/rmq-data/", "disk_type": "nas", "real_path": "/var/lib/docker/volumes/rmq-redis-etcd_nas-rmqbroker-163/_data"},
    {"type": "redis", "ip": "10.10.2.159", "purpose": "IM redis服务器", "data_dir": "/", "disk_type": "本地", "real_path": ""},
    {"type": "pgsql", "ip": "10.20.4.6", "purpose": "东莞mes数据库", "data_dir": "/volume1/PG/data1", "disk_type": "nas", "real_path": "/var/lib/docker/volumes/pg_sc_data/_data"},
    {"type": "mssql", "ip": "10.20.4.7", "purpose": "东莞mes数据库", "data_dir": "E盘", "disk_type": "本地", "real_path": ""},
]

ZABBIX_CONFIG = {
    "sz": {
        "url": "http://10.10.0.22/zabbix/api_jsonrpc.php",
        "user": "Admin",
        "password": "zabbix"
    },
    "dg": {
        "url": "http://10.20.4.10/zabbix/api_jsonrpc.php",
        "user": "Admin",
        "password": "zabbix"
    }
}

SSH_USER = "nas_monitor"
SSH_PASS = "Aa_123456"

# 特殊节点的磁盘监控项前缀映射
SPECIAL_DISK_PREFIX = {
    "10.20.4.7": "vfs.fs.dependent.size",  # 东莞mes数据库
}
DEFAULT_DISK_PREFIX = "vfs.fs.size"

# ==================== 邮件发送配置 ====================
TOKEN_API = "http://test.7ieis.com/api/identity/Account/Simulate"
TOKEN_CODE = "10004"
AUTH_HEADER = "Basic VE9PTDpscUAjMTIz"  # 固定的 Base64 凭证

UPLOAD_API = "http://test.7ieis.com/api/object-storage/FileUpDown/UploadFileByStream"
MAIL_API = "http://test.7ieis.com/api/email/AppMails/LQ_Mail_SaveSystemMail"
NOTIFY_API = "http://test.7ieis.com/api/im-server/MessageExchange/NotifyEmail"

TENANT_ID = 16
MAIL_RECIPIENTS = "E69819,E16307"
MAIL_CC = "E69776"                          # 新增抄送人
NOTIFY_RECIPIENT_IDS = ["69819", "16307", "69776"]  # 新增通知收件人 69776

# ==================== Zabbix API客户端 ====================
class ZabbixClient:
    def __init__(self, url: str, user: str, password: str):
        self.url = url
        self.auth = None
        self.login(user, password)

    def login(self, user: str, password: str):
        data = {
            "jsonrpc": "2.0",
            "method": "user.login",
            "params": {
                "username": user,
                "password": password
            },
            "id": 1
        }
        resp = requests.post(self.url, json=data, timeout=10)
        resp.raise_for_status()
        result = resp.json()
        if 'result' in result:
            self.auth = result['result']
        else:
            raise Exception(f"Zabbix登录失败: {result.get('error')}")

    def call(self, method: str, params: dict) -> dict:
        data = {
            "jsonrpc": "2.0",
            "method": method,
            "params": params,
            "auth": self.auth,
            "id": 1
        }
        resp = requests.post(self.url, json=data, timeout=10)
        resp.raise_for_status()
        return resp.json()

    def get_host_id(self, ip: str) -> Optional[str]:
        result = self.call("host.get", {
            "filter": {"ip": ip},
            "output": ["hostid"]
        })
        if result.get('result'):
            return result['result'][0]['hostid']
        result = self.call("host.get", {
            "filter": {"host": ip},
            "output": ["hostid"]
        })
        if result.get('result'):
            return result['result'][0]['hostid']
        return None

    def get_item_value(self, host_id: str, key_pattern: str) -> Optional[str]:
        result = self.call("item.get", {
            "hostids": host_id,
            "search": {"key_": key_pattern},
            "output": ["itemid", "key_", "lastvalue", "name"],
            "sortfield": "itemid",
            "limit": 1
        })
        if result.get('result'):
            return result['result'][0]['lastvalue']
        return None

# ==================== SSH工具（禁用公钥认证）====================
def ssh_exec(host: str, command: str) -> Optional[str]:
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        # 禁用公钥认证和 ssh-agent，只使用密码认证
        client.connect(
            host,
            username=SSH_USER,
            password=SSH_PASS,
            timeout=15,
            look_for_keys=False,   # 不尝试加载私钥
            allow_agent=False       # 不尝试使用 ssh-agent
        )
        stdin, stdout, stderr = client.exec_command(command)
        output = stdout.read().decode('utf-8').strip()
        error = stderr.read().decode('utf-8').strip()
        if error:
            print(f"[SSH] {host} 错误: {error}")
        return output
    except Exception as e:
        print(f"[SSH] {host} 连接失败: {e}")
        return None
    finally:
        client.close()

def get_linux_disk_usage_ssh(host: str, path: str) -> Dict[str, str]:
    cmd = f"sudo df -h {path} | awk 'NR==2 {{print $5, $2, $3, $4}}'"
    output = ssh_exec(host, cmd)
    if output:
        parts = output.split()
        if len(parts) >= 4:
            return {
                "usage": parts[0],
                "total": parts[1],
                "used": parts[2],
                "free": parts[3]
            }
    return {}

# ==================== 工具函数 ====================
def _format_bytes(bytes_val: float) -> str:
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if bytes_val < 1024.0:
            return f"{bytes_val:.2f}{unit}"
        bytes_val /= 1024.0
    return f"{bytes_val:.2f}PB"

# ==================== 巡检主逻辑 ====================
def run_inspection() -> str:
    """
    执行巡检，生成 Excel 报告
    返回生成的 Excel 文件路径
    """
    zabbix_sz = ZabbixClient(**ZABBIX_CONFIG["sz"])
    zabbix_dg = ZabbixClient(**ZABBIX_CONFIG["dg"])

    results = []

    for node in NODES:
        ip = node["ip"]
        node_type = node["type"]
        purpose = node["purpose"]
        data_dir = node["data_dir"]
        disk_type = node["disk_type"]
        real_path = node.get("real_path", "")

        print(f"正在处理 {ip} ({purpose})...")

        if ip.startswith("10.10."):
            zapi = zabbix_sz
        elif ip.startswith("10.20."):
            zapi = zabbix_dg
        else:
            print(f"未知IP段: {ip}")
            continue

        host_id = zapi.get_host_id(ip)
        if not host_id:
            print(f"⚠️ 在Zabbix中未找到主机 {ip}，无法获取CPU/内存数据")
            cpu = mem = None
        else:
            if node_type in ("pgsql", "redis", "rocketmq"):  # Linux
                cpu_key = "system.cpu.util[,idle]"
                cpu_idle = zapi.get_item_value(host_id, cpu_key)
                if cpu_idle:
                    cpu = 100 - float(cpu_idle)
                else:
                    print(f"⚠️ Zabbix中未找到 {ip} 的CPU使用率监控项: {cpu_key}")
                    cpu = None
                mem_key = "vm.memory.utilization"
                mem_item = zapi.get_item_value(host_id, mem_key)
                if mem_item:
                    mem = float(mem_item)
                else:
                    print(f"⚠️ Zabbix中未找到 {ip} 的内存使用率监控项: {mem_key}")
                    mem = None
            else:  # Windows
                cpu_key = "system.cpu.util"
                cpu_item = zapi.get_item_value(host_id, cpu_key)
                if cpu_item:
                    cpu = float(cpu_item)
                else:
                    print(f"⚠️ Zabbix中未找到 {ip} 的CPU使用率监控项: {cpu_key}")
                    cpu = None
                mem_key = "vm.memory.util"
                mem_item = zapi.get_item_value(host_id, mem_key)
                if mem_item:
                    mem = float(mem_item)
                else:
                    print(f"⚠️ Zabbix中未找到 {ip} 的内存使用率监控项: {mem_key}")
                    mem = None

        disk_usage = None
        disk_used = None

        if node_type in ("pgsql", "redis", "rocketmq") and real_path:
            ssh_disk = get_linux_disk_usage_ssh(ip, real_path)
            if ssh_disk:
                disk_usage = ssh_disk.get("usage")
                disk_used = ssh_disk.get("used")
            else:
                print(f"⚠️ SSH获取磁盘信息失败 {ip}:{real_path}")
        else:
            if host_id:
                disk_prefix = SPECIAL_DISK_PREFIX.get(ip, DEFAULT_DISK_PREFIX)
                if node_type == "mssql":
                    drive_letter = data_dir[0].upper() + ":"
                    usage_key = f"{disk_prefix}[{drive_letter},pused]"
                    usage_item = zapi.get_item_value(host_id, usage_key)
                    if usage_item:
                        disk_usage = f"{float(usage_item):.2f}%"
                    else:
                        print(f"⚠️ Zabbix中未找到 {ip} 的磁盘使用率监控项: {usage_key}")
                    used_key = f"{disk_prefix}[{drive_letter},used]"
                    used_item = zapi.get_item_value(host_id, used_key)
                    if used_item:
                        disk_used = _format_bytes(float(used_item))
                    else:
                        print(f"⚠️ Zabbix中未找到 {ip} 的磁盘已用监控项: {used_key}")
                else:
                    usage_key = f"{disk_prefix}[{data_dir},pused]"
                    usage_item = zapi.get_item_value(host_id, usage_key)
                    if usage_item:
                        disk_usage = f"{float(usage_item):.2f}%"
                    else:
                        print(f"⚠️ Zabbix中未找到 {ip} 的磁盘使用率监控项: {usage_key}")
                    used_key = f"{disk_prefix}[{data_dir},used]"
                    used_item = zapi.get_item_value(host_id, used_key)
                    if used_item:
                        disk_used = _format_bytes(float(used_item))
                    else:
                        print(f"⚠️ Zabbix中未找到 {ip} 的磁盘已用监控项: {used_key}")
            else:
                print(f"⚠️ 主机 {ip} 无Zabbix数据，无法获取磁盘信息")

        # 组装结果
        result = {
            "数据库/中间件类型": node_type,
            "IP地址": ip,
            "用途": purpose,
            "CPU使用率": f"{cpu:.2f}%" if cpu is not None else "N/A",
            "内存使用率": f"{mem:.2f}%" if mem is not None else "N/A",
            "数据大小": disk_used if disk_used is not None else "N/A",
            "数据盘使用率": disk_usage if disk_usage is not None else "N/A",
            "数据目录": data_dir,
            "磁盘类型": disk_type
        }
        results.append(result)
        time.sleep(0.5)

    # 转换为DataFrame
    df = pd.DataFrame(results)

    # 生成Excel报告
    wb = Workbook()
    ws = wb.active

    today_str = datetime.datetime.now().strftime("%Y%m%d")
    sheet_name = f"巡检-{today_str}"
    ws.title = sheet_name

    # 标题
    title = f"巡检-{today_str}"
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列名
    columns = list(df.columns)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # 修改表头背景色为 #4caf50
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    # 写入数据
    for row_idx, row_data in enumerate(df.values, start=3):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 调整列宽
    column_widths = {
        'A': 18,
        'B': 15,
        'C': 25,
        'D': 12,
        'E': 12,
        'F': 12,
        'G': 15,
        'H': 30,
        'I': 12,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    excel_filename = f"数据库及中间件磁盘巡检{today_str}.xlsx"
    wb.save(excel_filename)
    print(f"\n✅ Excel报告已生成：{excel_filename}")

    # 控制台输出
    print("\n=== 监控数据汇总 ===")
    print(df.to_string(index=False))

    return excel_filename

# ==================== 邮件发送相关函数 ====================
def get_token():
    """使用固定的 Basic 凭证获取 Bearer Token"""
    headers = {"Authorization": AUTH_HEADER}
    params = {"code": TOKEN_CODE}
    try:
        resp = requests.get(TOKEN_API, params=params, headers=headers, timeout=10)
        resp.raise_for_status()
        logging.info("Token 响应内容: %s", resp.text)
        data = resp.json()
        if isinstance(data, dict):
            token = data.get("Data")
            if token:
                if isinstance(token, str):
                    return token.strip()
                elif isinstance(token, dict):
                    inner_token = token.get("Token") or token.get("token") or token.get("access_token")
                    if inner_token:
                        return inner_token
                    else:
                        return json.dumps(token)
                else:
                    return str(token)
            else:
                token = data.get("token") or data.get("access_token")
                if token:
                    return token
                raise Exception(f"响应中未找到 token 字段: {data}")
        else:
            return resp.text.strip()
    except Exception as e:
        logging.error("获取 Token 失败: %s", e)
        raise


def upload_file(token, file_path):
    """上传文件，返回 fileId"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "text/plain"
    }
    filename = os.path.basename(file_path)
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    with open(file_path, "rb") as f:
        files = {"file": (filename, f, mime_type)}
        try:
            resp = requests.post(UPLOAD_API, headers=headers, files=files, timeout=30)
            resp.raise_for_status()
            try:
                result = resp.json()
                if isinstance(result, dict):
                    file_id = result.get("Data") or result.get("fileId") or result.get("id")
                    if file_id:
                        return file_id
                    else:
                        logging.warning("未知的 JSON 结构，返回完整响应: %s", result)
                        return result
                else:
                    return result
            except:
                return resp.text.strip()
        except Exception as e:
            logging.error("文件上传失败: %s", e)
            raise


def excel_to_html_table(file_path, report_date):

    try:
        df_raw = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', header=None)

        title = f"巡检-{report_date}"

        headers_raw = df_raw.iloc[1, :].tolist()
        headers = [str(h).strip() if not pd.isna(h) else '' for h in headers_raw]

        data_raw = df_raw.iloc[2:, :]

        df = pd.DataFrame(data_raw.values, columns=headers)
        df = df.dropna(axis=1, how='all')

        columns = df.columns.tolist()
        col_count = len(columns)

        styles = """
        <style>
        table.report-table{
            border-collapse:collapse !important;
            border-spacing:0 !important;
        }
        table.report-table th{
            border:1px solid #000000 !important;
            background:#4caf50;
            color:black;
            padding:8px;
        }
        table.report-table td{
            border:1px solid #000000 !important;
            padding:8px;
            background:#ffffff;
        }
        </style>
        """

        html_parts = [styles]

        html_parts.append(
        f'''
<table class="report-table"
border="1"
bordercolor="#000000"
cellspacing="0"
cellpadding="0"
style="border-collapse:collapse;border-spacing:0;border:1px solid #000000;width:100%;">
'''
        )

        html_parts.append(
            f'<tr>'
            f'<th colspan="{col_count}" '
            f'style="border:1px solid #000000 !important;'
            f'font-size:16px;'
            f'text-align:center;'
            f'background:#ffffff !important;">'
            f'{title}</th>'
            f'</tr>'
        )

        html_parts.append('<tr>')

        for col in columns:
            html_parts.append(
                f'<th style="border:1px solid #000000 !important;">{col}</th>'
            )

        html_parts.append('</tr>')

        for _, row in df.iterrows():

            html_parts.append('<tr>')

            for col in columns:

                val = row[col]
                if pd.isna(val):
                    val = ''

                html_parts.append(
                    f'<td style="border:1px solid #000000 !important;">{val}</td>'
                )

            html_parts.append('</tr>')

        html_parts.append('</table>')

        return ''.join(html_parts)

    except Exception as e:
        logging.error("Excel 转 HTML 失败: %s", e)
        return None


def send_mail_with_attachment(excel_path):
    """
    使用生成的 Excel 文件发送邮件
    """
    # 获取 Token
    try:
        token = get_token()
        logging.info("Token 获取成功")
    except Exception as e:
        logging.error("获取 Token 失败: %s", e)
        return False

    # 上传文件获取 fileId
    try:
        file_id = upload_file(token, excel_path)
        logging.info(f"附件上传成功，fileId: {file_id}")
    except Exception as e:
        logging.error("文件上传失败: %s", e)
        return False

    # 生成 HTML 表格（使用当前日期）
    today_str = datetime.datetime.now().strftime("%Y%m%d")
    html_table = excel_to_html_table(excel_path, today_str)

    # 构造邮件内容
    subject = f"{today_str}-巡检"
    body_intro = "Hi， all:<br>以下是今天的巡检情况，请查收！"

    if html_table:
        body = f"{body_intro}<br><br>{html_table}"
    else:
        body = body_intro + "<br><br>（注：无法生成数据表格，请直接查看附件。）"

    attachments_str = json.dumps([file_id])

    # 新增 CCUserIds 字段
    mail_data = {
        "Subject": subject,
        "RecipientIds": MAIL_RECIPIENTS,
        "CCUserIds": MAIL_CC,                # 抄送人
        "Body": body,
        "Attachments": attachments_str
    }
    mail_payload = {
        "TenantId": TENANT_ID,
        "Data": json.dumps(mail_data, ensure_ascii=False)
    }

    # 发送邮件
    try:
        mail_resp = requests.post(MAIL_API, json=mail_payload, timeout=5)
        mail_resp.raise_for_status()
        logging.info("[MAIL] 状态码=%s 响应=%s", mail_resp.status_code, mail_resp.text)
    except requests.exceptions.RequestException as e:
        logging.error("[MAIL] 请求失败: %s", e)
        return False

    # 发送通知（使用更新后的收件人列表）
    notify_payload = {
        "Type": 0,
        "EnterpriseId": TENANT_ID,
        "RecipientIds": NOTIFY_RECIPIENT_IDS
    }
    try:
        notify_resp = requests.post(NOTIFY_API, json=notify_payload, timeout=5)
        notify_resp.raise_for_status()
        logging.info("[PUSH] 通知发送成功")
    except requests.exceptions.RequestException as e:
        logging.error("[PUSH] 通知发送失败: %s", e)

    return True


# ==================== 主函数 ====================
def main():
    print("="*50)
    print("开始执行巡检...")
    excel_file = run_inspection()
    print("="*50)
    print("开始发送邮件...")
    success = send_mail_with_attachment(excel_file)
    if success:
        print("✅ 邮件发送成功！")
    else:
        print("❌ 邮件发送失败，请检查日志。")
    print("="*50)


if __name__ == "__main__":
    main()
